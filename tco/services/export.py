"""Экспорт данных в CSV и Excel (SCOPE-R P §16).

Экспорт всегда соответствует выбранному фильтру и включает поля качества,
статуса, источников и версии методики. Текстовые значения проходят защиту
от CSV injection (SCOPE-R E §7).
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from datetime import date, timedelta
from typing import Any, Sequence

from sqlalchemy import select
from sqlalchemy.orm import Session

from tco.core.config import get_settings
from tco.core.enums import ExportDataset, ExportFormat
from tco.core.errors import ValidationError
from tco.core.logging import get_logger
from tco.core.utils import csv_safe, round_display, sha256_bytes, utcnow
from tco.db.models.job import ExportArtifact, Job
from tco.db.models.offer import Offer
from tco.db.models.run import ScenarioRun
from tco.db.models.scenario import TravelScenario
from tco.db.models.source import SourceMetric
from tco.services.dashboard import DashboardFilters, _apply_filters
from tco.storage.raw_store import RawStore, get_raw_store
from tco.version import METRIC_DISCLAIMER_RU, METRIC_TITLE_RU, version_payload

logger = get_logger(__name__)

#: Колонки выгрузки ScenarioRun. Порядок фиксирован — на него опираются
#: потребители выгрузки.
RUN_COLUMNS: list[tuple[str, str]] = [
    ("scenario_code", "Код сценария"),
    ("scenario_name", "Название сценария"),
    ("origin", "Город отправления"),
    ("destination", "Город назначения"),
    ("departure_date", "Дата начала"),
    ("return_date", "Дата окончания"),
    ("nights", "Ночей"),
    ("adults", "Взрослых"),
    ("children_ages", "Возраст детей"),
    ("traveler_count", "Всего туристов"),
    ("transport_type", "Транспорт"),
    ("flight_fare_type", "Авиатариф"),
    ("rail_class", "Класс ЖД"),
    ("accommodation_type", "Тип размещения"),
    ("stars", "Звездность"),
    ("meal_type", "Питание"),
    ("cancellation_filter", "Условие отмены"),
    ("run_id", "ID расчета"),
    ("run_type", "Тип расчета"),
    ("observation_date", "Дата наблюдения"),
    ("run_date", "Дата и время расчета"),
    ("lead_time_days", "Горизонт бронирования, дней"),
    ("status", "Статус"),
    ("component_statuses", "Компонентные статусы"),
    ("transport_p25", "Транспорт P25"),
    ("transport_median", "Транспорт медиана"),
    ("transport_p75", "Транспорт P75"),
    ("hotel_p25", "Проживание P25"),
    ("hotel_median", "Проживание медиана"),
    ("hotel_p75", "Проживание P75"),
    ("total_p25", "Итого P25"),
    ("total_estimated_cost", METRIC_TITLE_RU),
    ("total_p75", "Итого P75"),
    ("price_per_person", "Стоимость на человека"),
    ("transport_share", "Доля транспорта"),
    ("currency", "Валюта"),
    ("quality_score", "Quality Score"),
    ("confidence_level", "Уверенность"),
    ("confidence_reason", "Причина уровня уверенности"),
    ("source_count", "Число источников"),
    ("source_codes", "Источники"),
    ("transport_source_count", "Источников транспорта"),
    ("hotel_source_count", "Источников проживания"),
    ("transport_disagreement", "Расхождение по транспорту"),
    ("hotel_disagreement", "Расхождение по проживанию"),
    ("valid_offer_count", "Учтено предложений"),
    ("excluded_offer_count", "Исключено предложений"),
    ("outlier_offer_count", "Выбросов"),
    ("profile_code", "Профиль"),
    ("profile_version", "Версия профиля"),
    ("engine_version", "Версия движка"),
    ("normalization_version", "Версия нормализации"),
    ("market_snapshot_id", "ID снимка рынка"),
    ("served_from_cache", "Из кэша"),
    ("contains_synthetic_data", "Синтетические данные"),
]

OFFER_COLUMNS: list[tuple[str, str]] = [
    ("offer_id", "ID предложения"),
    ("market_snapshot_id", "ID снимка"),
    ("scenario_code", "Код сценария"),
    ("source_code", "Источник"),
    ("offer_type", "Тип предложения"),
    ("collected_at", "Собрано"),
    ("currency", "Валюта"),
    ("total_price", "Цена"),
    ("price_per_night", "Цена за ночь"),
    ("validity_status", "Валидность"),
    ("classification_status", "Классификация"),
    ("is_duplicate", "Дубликат"),
    ("is_outlier", "Выброс"),
    ("matches_profile", "Соответствует профилю"),
    ("exclusion_reason", "Причина исключения"),
    ("exclusion_detail", "Детали исключения"),
    ("detail", "Описание"),
]

SOURCE_METRIC_COLUMNS: list[tuple[str, str]] = [
    ("source_code", "Источник"),
    ("offer_type", "Тип предложения"),
    ("observed_at", "Момент наблюдения"),
    ("outcome", "Итог"),
    ("latency_ms", "Задержка, мс"),
    ("attempts", "Попыток"),
    ("raw_offer_count", "Сырых предложений"),
    ("normalized_offer_count", "Нормализовано"),
    ("valid_offer_count", "Валидных"),
    ("invalid_offer_count", "Невалидных"),
    ("unclassified_offer_count", "Неклассифицированных"),
    ("required_field_completeness", "Полнота полей"),
    ("error_code", "Код ошибки"),
    ("error_message", "Сообщение об ошибке"),
]


@dataclass(slots=True)
class ExportRequest:
    dataset: ExportDataset
    export_format: ExportFormat
    filters: DashboardFilters
    limit: int = 100_000
    snapshot_id: str | None = None
    window_days: int = 30

    def as_dict(self) -> dict[str, Any]:
        return {
            "dataset": self.dataset.value,
            "format": self.export_format.value,
            "filters": self.filters.as_dict(),
            "limit": self.limit,
            "snapshot_id": self.snapshot_id,
            "window_days": self.window_days,
        }


# --------------------------------------------------------------------------- #
# Сбор строк
# --------------------------------------------------------------------------- #


def collect_run_rows(session: Session, request: ExportRequest) -> list[dict[str, Any]]:
    stmt = select(ScenarioRun)
    stmt = _apply_filters(stmt, request.filters).order_by(
        ScenarioRun.observation_date.desc(), ScenarioRun.started_at.desc()
    )
    runs = session.scalars(stmt.limit(request.limit)).all()

    rows: list[dict[str, Any]] = []
    for run in runs:
        scenario: TravelScenario = run.scenario
        rows.append(
            {
                "scenario_code": scenario.code,
                "scenario_name": scenario.name,
                "origin": scenario.origin_city.name,
                "destination": scenario.destination_city.name,
                "departure_date": scenario.departure_date.isoformat(),
                "return_date": scenario.return_date.isoformat(),
                "nights": scenario.nights,
                "adults": scenario.adults,
                "children_ages": ";".join(str(age) for age in (scenario.children_ages or [])),
                "traveler_count": run.traveler_count,
                "transport_type": scenario.transport_type,
                "flight_fare_type": scenario.flight_fare_type or "",
                "rail_class": scenario.rail_class or "",
                "accommodation_type": scenario.accommodation_type,
                "stars": scenario.stars,
                "meal_type": scenario.meal_type,
                "cancellation_filter": scenario.cancellation_filter,
                "run_id": str(run.id),
                "run_type": run.run_type,
                "observation_date": run.observation_date.isoformat(),
                "run_date": run.started_at.isoformat(),
                "lead_time_days": run.lead_time_days,
                "status": run.status,
                "component_statuses": ";".join(run.component_statuses or []),
                "transport_p25": round_display(run.transport_p25, 0),
                "transport_median": round_display(run.transport_median, 0),
                "transport_p75": round_display(run.transport_p75, 0),
                "hotel_p25": round_display(run.hotel_p25, 0),
                "hotel_median": round_display(run.hotel_median, 0),
                "hotel_p75": round_display(run.hotel_p75, 0),
                "total_p25": round_display(run.total_p25, 0),
                "total_estimated_cost": round_display(run.total_estimated_cost, 0),
                "total_p75": round_display(run.total_p75, 0),
                "price_per_person": round_display(run.price_per_person, 0),
                "transport_share": run.transport_share,
                "currency": run.currency,
                "quality_score": run.quality_score,
                "confidence_level": run.confidence_level or "",
                "confidence_reason": run.confidence_reason or "",
                "source_count": run.source_count,
                "source_codes": ";".join(run.source_codes or []),
                "transport_source_count": run.transport_source_count,
                "hotel_source_count": run.hotel_source_count,
                "transport_disagreement": run.transport_disagreement,
                "hotel_disagreement": run.hotel_disagreement,
                "valid_offer_count": run.valid_offer_count,
                "excluded_offer_count": run.excluded_offer_count,
                "outlier_offer_count": run.outlier_offer_count,
                "profile_code": run.profile_code,
                "profile_version": run.profile_version,
                "engine_version": run.engine_version,
                "normalization_version": run.normalization_version,
                "market_snapshot_id": str(run.market_snapshot_id) if run.market_snapshot_id else "",
                "served_from_cache": run.served_from_cache,
                "contains_synthetic_data": run.contains_synthetic_data,
            }
        )
    return rows


def collect_offer_rows(session: Session, request: ExportRequest) -> list[dict[str, Any]]:
    if not request.snapshot_id:
        raise ValidationError("Для выгрузки предложений требуется snapshot_id")
    import uuid as uuid_lib

    offers = session.scalars(
        select(Offer)
        .where(Offer.market_snapshot_id == uuid_lib.UUID(request.snapshot_id))
        .order_by(Offer.offer_type, Offer.source_code, Offer.total_price)
        .limit(request.limit)
    ).all()

    rows: list[dict[str, Any]] = []
    for offer in offers:
        scenario_code = offer.scenario.code if getattr(offer, "scenario", None) else ""
        rows.append(
            {
                "offer_id": str(offer.id),
                "market_snapshot_id": str(offer.market_snapshot_id),
                "scenario_code": scenario_code,
                "source_code": offer.source_code,
                "offer_type": offer.offer_type,
                "collected_at": offer.collected_at.isoformat(),
                "currency": offer.currency,
                "total_price": round_display(offer.total_price, 2),
                "price_per_night": round_display(offer.price_per_night, 2),
                "validity_status": offer.validity_status,
                "classification_status": offer.classification_status,
                "is_duplicate": offer.is_duplicate,
                "is_outlier": offer.is_outlier,
                "matches_profile": offer.matches_profile,
                "exclusion_reason": offer.exclusion_reason,
                "exclusion_detail": offer.exclusion_detail or "",
                "detail": _offer_description(offer),
            }
        )
    return rows


def _offer_description(offer: Offer) -> str:
    if offer.flight is not None:
        carriers = ", ".join(offer.flight.marketing_carriers or [])
        return (
            f"{carriers} · пересадок {offer.flight.outbound_stops}/{offer.flight.inbound_stops} · "
            f"багаж {offer.flight.baggage_type}"
        )
    if offer.rail is not None:
        return (
            f"Поезда {offer.rail.outbound_train_number}/{offer.rail.inbound_train_number} · "
            f"{offer.rail.car_type}"
        )
    if offer.accommodation is not None:
        stars = offer.accommodation.stars or offer.accommodation.stars_status
        return (
            f"{offer.accommodation.property_name} ({stars}) · "
            f"{offer.accommodation.room_name} · питание {offer.accommodation.meal_type}"
        )
    return ""


def collect_source_metric_rows(session: Session, request: ExportRequest) -> list[dict[str, Any]]:
    since = utcnow() - timedelta(days=request.window_days)
    metrics = session.scalars(
        select(SourceMetric)
        .where(SourceMetric.observed_at >= since)
        .order_by(SourceMetric.observed_at.desc())
        .limit(request.limit)
    ).all()
    return [
        {
            "source_code": metric.source.code if metric.source else "",
            "offer_type": metric.offer_type,
            "observed_at": metric.observed_at.isoformat(),
            "outcome": metric.outcome,
            "latency_ms": metric.latency_ms,
            "attempts": metric.attempts,
            "raw_offer_count": metric.raw_offer_count,
            "normalized_offer_count": metric.normalized_offer_count,
            "valid_offer_count": metric.valid_offer_count,
            "invalid_offer_count": metric.invalid_offer_count,
            "unclassified_offer_count": metric.unclassified_offer_count,
            "required_field_completeness": metric.required_field_completeness,
            "error_code": metric.error_code or "",
            "error_message": metric.error_message or "",
        }
        for metric in metrics
    ]


COLUMNS_BY_DATASET = {
    ExportDataset.SCENARIO_RUNS: RUN_COLUMNS,
    ExportDataset.OFFERS: OFFER_COLUMNS,
    ExportDataset.SOURCE_METRICS: SOURCE_METRIC_COLUMNS,
}

COLLECTORS = {
    ExportDataset.SCENARIO_RUNS: collect_run_rows,
    ExportDataset.OFFERS: collect_offer_rows,
    ExportDataset.SOURCE_METRICS: collect_source_metric_rows,
}


# --------------------------------------------------------------------------- #
# Рендеринг
# --------------------------------------------------------------------------- #


def render_csv(rows: Sequence[dict[str, Any]], columns: Sequence[tuple[str, str]]) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=";", lineterminator="\r\n", quoting=csv.QUOTE_MINIMAL)
    writer.writerow([title for _, title in columns])
    for row in rows:
        writer.writerow([csv_safe(_cell(row.get(key))) for key, _ in columns])
    # BOM: Excel корректно открывает UTF-8 CSV только с ним.
    return b"\xef\xbb\xbf" + buffer.getvalue().encode("utf-8")


def render_xlsx(
    rows: Sequence[dict[str, Any]],
    columns: Sequence[tuple[str, str]],
    *,
    sheet_title: str,
    meta: dict[str, Any],
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_title[:31]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5597")
    for index, (_, title) in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=index, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.column_dimensions[get_column_letter(index)].width = min(max(len(title) + 4, 12), 42)
    sheet.freeze_panes = "A2"

    for row_index, row in enumerate(rows, start=2):
        for col_index, (key, _) in enumerate(columns, start=1):
            sheet.cell(row=row_index, column=col_index, value=csv_safe(_cell(row.get(key))))

    # Отдельный лист с параметрами выгрузки и оговоркой о смысле показателя.
    info = workbook.create_sheet("Параметры выгрузки")
    info.column_dimensions["A"].width = 34
    info.column_dimensions["B"].width = 90
    info_rows = [
        ("Показатель", METRIC_TITLE_RU),
        ("Оговорка", METRIC_DISCLAIMER_RU),
        ("Строк", len(rows)),
        *[(key, _cell(value)) for key, value in meta.items()],
    ]
    for row_index, (key, value) in enumerate(info_rows, start=1):
        info.cell(row=row_index, column=1, value=key).font = Font(bold=True)
        cell = info.cell(row=row_index, column=2, value=csv_safe(value))
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "да" if value else "нет"
    if isinstance(value, (list, tuple)):
        return ";".join(str(item) for item in value)
    if isinstance(value, dict):
        import json

        return json.dumps(value, ensure_ascii=False)
    return value


# --------------------------------------------------------------------------- #
# Выполнение
# --------------------------------------------------------------------------- #


def build_export(
    session: Session, request: ExportRequest
) -> tuple[bytes, str, int, str]:
    """Готовит выгрузку. Возвращает ``(данные, имя_файла, строк, content_type)``."""
    columns = COLUMNS_BY_DATASET[request.dataset]
    rows = COLLECTORS[request.dataset](session, request)
    stamp = utcnow().strftime("%Y%m%d-%H%M%S")
    meta = {
        "Набор данных": request.dataset.value,
        "Фильтры": request.filters.as_dict(),
        "Сформировано": utcnow().isoformat(),
        **{f"Версия: {key}": value for key, value in version_payload().items()},
    }

    if request.export_format == ExportFormat.CSV:
        return (
            render_csv(rows, columns),
            f"{request.dataset.value.lower()}-{stamp}.csv",
            len(rows),
            "text/csv; charset=utf-8",
        )
    return (
        render_xlsx(rows, columns, sheet_title=request.dataset.value, meta=meta),
        f"{request.dataset.value.lower()}-{stamp}.xlsx",
        len(rows),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def request_from_params(params: dict[str, Any]) -> ExportRequest:
    """Восстанавливает запрос экспорта из параметров задачи."""
    filters_raw = params.get("filters") or {}
    from tco.core.enums import TransportType

    transport = filters_raw.get("transport_type")
    filters = DashboardFilters(
        origin_city_code=filters_raw.get("origin_city_code"),
        destination_city_code=filters_raw.get("destination_city_code"),
        transport_type=TransportType(transport) if transport else None,
        accommodation_type=filters_raw.get("accommodation_type"),
        stars=filters_raw.get("stars"),
        scenario_codes=filters_raw.get("scenario_codes") or (),
        date_from=date.fromisoformat(filters_raw["date_from"])
        if filters_raw.get("date_from")
        else None,
        date_to=date.fromisoformat(filters_raw["date_to"]) if filters_raw.get("date_to") else None,
        profile_version=filters_raw.get("profile_version"),
        include_synthetic=bool(filters_raw.get("include_synthetic", False)),
        complete_only=bool(filters_raw.get("complete_only", False)),
        min_quality_score=filters_raw.get("min_quality_score"),
    )
    return ExportRequest(
        dataset=ExportDataset(params.get("dataset", ExportDataset.SCENARIO_RUNS.value)),
        export_format=ExportFormat(params.get("format", ExportFormat.CSV.value)),
        filters=filters,
        limit=int(params.get("limit", 100_000)),
        snapshot_id=params.get("snapshot_id"),
        window_days=int(params.get("window_days", 30)),
    )


def run_export_job(
    session: Session, job: Job, raw_store: RawStore | None = None
) -> ExportArtifact:
    """Выполняет задачу экспорта и сохраняет артефакт."""
    settings = get_settings()
    raw_store = raw_store or get_raw_store()
    request = request_from_params(dict(job.params or {}))
    data, filename, row_count, _content_type = build_export(session, request)

    stored = raw_store.put_bytes(
        data,
        source_code="export",
        scenario_code=request.dataset.value.lower(),
        request_id=str(job.id),
        content_type="application/octet-stream",
        extension=f"{filename.rsplit('.', 1)[-1]}.gz",
        kind="export",
    )
    artifact = ExportArtifact(
        job_id=job.id,
        dataset=request.dataset.value,
        export_format=request.export_format.value,
        filename=filename,
        storage_ref=stored.ref,
        row_count=row_count,
        size_bytes=len(data),
        checksum_sha256=sha256_bytes(data),
        filters=request.filters.as_dict(),
        created_by=job.created_by,
        created_at=utcnow(),
        expires_at=utcnow() + timedelta(days=settings.retention_export_days)
        if settings.retention_export_days > 0
        else None,
    )
    session.add(artifact)
    session.flush()
    logger.info(
        "Экспорт сформирован",
        dataset=request.dataset.value,
        format=request.export_format.value,
        rows=row_count,
        size_bytes=len(data),
    )
    return artifact
